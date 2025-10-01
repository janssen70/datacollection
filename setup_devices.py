#!/usr/bin/env python3
import argparse
import ipaddress
import re
from openpyxl import load_workbook
import socket
import sys

from axis_device_tool.device_tool import VapixClient, WebAccess, StandardSSLContext

def get_local_ip():
    """Return the primary IPv4 address of the local machine."""
    try:
        # Doesn't actually send packets, just figures out routing table choice
        with socket.socket(socket.AF_INET, socket.SOCK_DGRAM) as s:
            s.connect(("8.8.8.8", 80))  # Google's DNS (arbitrary reachable IP)
            return s.getsockname()[0]
    except Exception as e:
        raise RuntimeError(f"Could not determine local IP: {e}")

def is_valid_ip_or_hostname(value: str) -> bool:
    """
    Check if value is a valid IP address or hostname.
    """
    try:
        ipaddress.ip_address(value)
        return True
    except ValueError:
        # Check for valid hostname (RFC 1035 simplified)
        hostname_regex = re.compile(
            r'^(?=.{1,253}$)(?!-)[A-Za-z0-9-]{1,63}(?<!-)(\.(?!-)[A-Za-z0-9-]{1,63}(?<!-))*$'
        )
        return bool(hostname_regex.match(value))

def is_valid_username(value: str) -> bool:
    """
    Check if username is alphanumeric or underscore only.
    """
    return bool(re.fullmatch(r'[A-Za-z0-9_]+', value))

def read_excel(filename: str):
    wb = load_workbook(filename=filename, read_only=True)
    ws = wb.active  # First sheet

    data = []
    first_row = True
    for row in ws.iter_rows(values_only=True):
        if first_row:
            first_row = False
            continue  # skip header row

        ip, username, password = row

        if not is_valid_ip_or_hostname(str(ip)):
            print(f'Invalid IP/hostname: {ip}', file=sys.stderr)
            continue

        if not is_valid_username(str(username)):
            print(f'Invalid username: {username}', file=sys.stderr)
            continue

        data.append({
            'ip': str(ip),
            'username': str(username),
            'password': str(password),
        })

    return data

def main():
   parser = argparse.ArgumentParser(description='Read Excel file into list of dicts.')
   parser.add_argument('-f', '--filename', help='Excel .xlsx file to read')
   parser.add_argument('-b', '--broker', help='Broker address')
   parser.add_argument('-t', '--topic', action = 'append', help = 'Event topic to publish on broker')
   # Todo: broker credentials
   args = parser.parse_args()

   if not (args.filename and args.broker):
       print(f'Please specify filename and broker address')
       exit(-1)

   for host in read_excel(args.filename):
      print(f'Setup {host["ip"]}...')
      w = WebAccess(host['ip'], context = StandardSSLContext())
      w.add_credentials(host['username'], host['password'])
      client = VapixClient(w)
      print(client.MQTTConfig(broker_addr = args.broker))
      publications = client.MQTTGetEventPublications()
      for topic in args.topic:
         publications.append(topic)
      print(client.MQTTAddEventPublications(publications))
      print(client.MQTTActivate())

if __name__ == '__main__':
   main()

#  vim: set nowrap sw=3 sts=3 et fdm=marker:
